"""
Cria o template Excel para input manual de papéis não disponíveis na API.

Estrutura do Excel:
  Aba "Cadastro": metadados do papel
  Aba "Fluxo":   fluxo de caixa (data, tipo, % ou valor absoluto)
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = BASE_DIR / "manual" / "template_fluxo.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
INPUT_FILL = PatternFill("solid", fgColor="D9E1F2")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _set_header(cell, value):
    cell.value = value
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _set_input(cell, value=None, comment: str = ""):
    if value is not None:
        cell.value = value
    cell.fill = INPUT_FILL
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")


def create_template():
    wb = openpyxl.Workbook()

    # ── Aba 1: Cadastro ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Cadastro"

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 40

    fields = [
        ("Ticker / Identificador", "MINHA_DEB_001", "Código único do papel (sem espaços)"),
        ("Descrição", "Debênture XYZ - Série 1", "Nome legível do papel"),
        ("Tipo", "DEB", "DEB | CRI | CRA | LCI | LCA | CDB | OUTRO"),
        ("Indexador", "IPCA", "IPCA | CDI | PRE | IGPM | IPCA+CDI"),
        ("% do Indexador (se CDI)", "100", "Ex: 100 = 100% CDI, 95.5 = 95,5% CDI (só para CDI)"),
        ("Spread / Taxa (% a.a.)", "3.50", "Spread sobre indexador ou taxa PRE"),
        ("Data de Início da Rentabilidade", "2024-01-15", "YYYY-MM-DD (data-base do VNA)"),
        ("VNA / VNE Inicial", "1000.00", "Valor nominal na data de início"),
        ("Data de Vencimento", "2030-01-15", "YYYY-MM-DD"),
    ]

    ws1["A1"] = "CADASTRO DO PAPEL"
    ws1["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws1.merge_cells("A1:C1")
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1["A2"] = "Campo"
    ws1["B2"] = "Valor"
    ws1["C2"] = "Instruções"
    for cell in [ws1["A2"], ws1["B2"], ws1["C2"]]:
        _set_header(cell, cell.value)

    for i, (campo, exemplo, instrucao) in enumerate(fields, start=3):
        ws1[f"A{i}"] = campo
        ws1[f"A{i}"].font = Font(bold=True, size=10)
        ws1[f"A{i}"].border = BORDER
        _set_input(ws1[f"B{i}"], exemplo)
        ws1[f"C{i}"] = instrucao
        ws1[f"C{i}"].font = Font(italic=True, color="595959", size=9)
        ws1[f"C{i}"].border = BORDER

    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 20

    # ── Aba 2: Fluxo ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Fluxo")

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 40

    ws2["A1"] = "FLUXO DE CAIXA"
    ws2["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws2.merge_cells("A1:E1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    headers_fluxo = [
        ("A2", "Data"),
        ("B2", "Tipo"),
        ("C2", "% do VNA"),
        ("D2", "Valor Absoluto"),
        ("E2", "Observação"),
    ]
    for cell_ref, val in headers_fluxo:
        _set_header(ws2[cell_ref], val)

    # Exemplos de linhas
    exemplos = [
        ("2025-01-15", "J", "3.500%", "", "Juros = spread × VNA × período"),
        ("2025-01-15", "A", "10.000%", "", "Amortização 10% do VNA nesta data"),
        ("2025-07-15", "J", "3.500%", "", ""),
        ("2025-07-15", "A", "10.000%", "", ""),
        ("2026-01-15", "J", "", "350.00", "Valor absoluto (quando % não se aplica)"),
        ("2026-01-15", "A", "", "1000.00", "Amortização final"),
    ]

    for i, (data, tipo, pct, valor, obs) in enumerate(exemplos, start=3):
        _set_input(ws2[f"A{i}"], data)
        _set_input(ws2[f"B{i}"], tipo)
        _set_input(ws2[f"C{i}"], pct)
        _set_input(ws2[f"D{i}"], valor if valor else None)
        ws2[f"E{i}"] = obs
        ws2[f"E{i}"].font = Font(italic=True, color="595959", size=9)
        ws2[f"E{i}"].border = BORDER

    ws2["A2"].comment = None
    ws2["B2"].comment = None

    # Nota de rodapé
    nota_row = len(exemplos) + 5
    ws2[f"A{nota_row}"] = "INSTRUÇÕES:"
    ws2[f"A{nota_row}"].font = Font(bold=True, color="1F4E79")
    notas = [
        "• Tipo: J = Juros (cupom)  |  A = Amortização",
        "• Preencha OU '% do VNA' OU 'Valor Absoluto', não os dois.",
        "• Datas no formato YYYY-MM-DD",
        "• Inclua todas as datas futuras. Datas passadas são ignoradas no cálculo.",
        "• Após preencher, use: python main.py manual carregar template_fluxo.xlsx",
    ]
    for j, nota in enumerate(notas, start=nota_row + 1):
        ws2[f"A{j}"] = nota
        ws2[f"A{j}"].font = Font(italic=True, size=9, color="595959")
        ws2.merge_cells(f"A{j}:E{j}")

    # ── Aba 3: Instruções Gerais ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Instruções")
    ws3["A1"] = "COMO USAR ESTE TEMPLATE"
    ws3["A1"].font = Font(bold=True, size=13, color="1F4E79")

    instrucoes = [
        "",
        "1. Aba 'Cadastro': preencha os dados do papel.",
        "2. Aba 'Fluxo': insira cada pagamento futuro em uma linha.",
        "   - Tipo J = Juros (cupom de juros)",
        "   - Tipo A = Amortização (devolução de principal)",
        "   - Use '% do VNA' quando o pagamento é definido como % do valor nominal.",
        "   - Use 'Valor Absoluto' quando o valor é fixo em R$.",
        "",
        "3. Salve o arquivo e execute:",
        "   python main.py manual carregar CAMINHO/template_fluxo.xlsx",
        "",
        "4. Após carregar, calcule normalmente:",
        "   python main.py pu  TICKER YIELD",
        "   python main.py taxa TICKER PU",
        "",
        "DICAS:",
        "• Para papéis CDI, os cupons futuros dependem da projeção da curva DI.",
        "  Nesse caso, prefira usar o sistema automático (API) quando possível.",
        "• Para papéis IPCA, preencha os % do VNA — o sistema aplica o IPCA automaticamente.",
        "• Para papéis PRE (pré-fixados), use Valor Absoluto para os cupons.",
    ]

    for i, linha in enumerate(instrucoes, start=2):
        ws3[f"A{i}"] = linha
        if linha.startswith(("1.", "2.", "3.", "4.", "DICAS")):
            ws3[f"A{i}"].font = Font(bold=True, size=10)
        else:
            ws3[f"A{i}"].font = Font(size=10)

    wb.save(TEMPLATE_PATH)
    print(f"Template criado: {TEMPLATE_PATH}")
    return str(TEMPLATE_PATH)


if __name__ == "__main__":
    create_template()
