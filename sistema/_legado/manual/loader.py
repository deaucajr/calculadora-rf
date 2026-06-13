"""
Carrega um template Excel preenchido e salva no banco de dados.
"""
import sys
from pathlib import Path
import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from src.db import get_conn


def load_from_excel(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)

    # Lê aba Cadastro
    ws1 = wb["Cadastro"]
    meta = {}
    for row in ws1.iter_rows(min_row=3, max_row=20, values_only=True):
        if row[0] and row[1]:
            meta[str(row[0]).strip()] = str(row[1]).strip() if row[1] else ""

    ticker = meta.get("Ticker / Identificador", "").upper().replace(" ", "_")
    if not ticker:
        raise ValueError("Ticker não preenchido na aba Cadastro.")

    tipo = meta.get("Tipo", "OUTRO")
    method = meta.get("Indexador", "PRE")
    startingdate = meta.get("Data de Início da Rentabilidade", "")
    expiredate = meta.get("Data de Vencimento", "")
    vna_inicial = float(meta.get("VNA / VNE Inicial", 1000) or 1000)
    yield_contract = float(meta.get("Spread / Taxa (% a.a.)", 0) or 0)
    descricao = meta.get("Descrição", ticker)

    # Lê aba Fluxo
    ws2 = wb["Fluxo"]
    flows = []
    for row in ws2.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            break
        event_date = str(row[0]).strip()
        event_type = str(row[1]).strip().upper() if row[1] else ""
        value_pct_raw = str(row[2]).strip().replace("%", "") if row[2] else ""
        value_abs_raw = str(row[3]).strip() if row[3] else ""

        if event_type not in ("J", "A"):
            continue
        if not event_date:
            continue

        value_pct = float(value_pct_raw) if value_pct_raw else None
        value_abs = float(value_abs_raw) if value_abs_raw else None

        if value_pct is None and value_abs is None:
            continue

        flows.append((ticker, event_date, event_type, value_pct, value_abs))

    if not flows:
        raise ValueError("Nenhum fluxo encontrado na aba 'Fluxo'.")

    with get_conn() as conn:
        conn.execute("""
            INSERT OR REPLACE INTO manual_bonds
              (ticker, descricao, tipo, method, startingdate, expiredate,
               vna_inicial, yield_contract)
            VALUES (?,?,?,?,?,?,?,?)
        """, (ticker, descricao, tipo, method, startingdate, expiredate,
              vna_inicial, yield_contract))

        conn.execute("DELETE FROM manual_flows WHERE ticker=?", (ticker,))
        conn.executemany("""
            INSERT INTO manual_flows
              (ticker, event_date, event_type, value_pct, value_abs)
            VALUES (?,?,?,?,?)
        """, flows)

    print(f"✓ Ativo '{ticker}' carregado: {len(flows)} fluxos")
    print(f"  Tipo: {tipo} | Indexador: {method} | Venc: {expiredate}")
    print(f"  Calcule com: python main.py pu {ticker} YIELD")
    return ticker


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python manual/loader.py CAMINHO/template_fluxo.xlsx")
        sys.exit(1)
    load_from_excel(sys.argv[1])
