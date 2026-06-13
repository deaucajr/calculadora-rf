"""
Gera o dashboard Excel completo com:
  - Carteira    : todos os ativos com PU/taxa/duration pré-calculados
  - Calculadora : calculadora interativa (via xlwings UDFs no Excel)
  - Cadastrar   : formulário de cadastro com campos auto-preenchíveis
  - Dados       : histórico de IPCA e CDI
"""
import sys
from pathlib import Path
from datetime import date as dt_date

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers, GradientFill,
)
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = BASE_DIR / "data" / "dashboard.xlsx"

# ── Paleta de cores ──────────────────────────────────────────────────────────
C_AZUL_ESCURO  = "1F3864"
C_AZUL_MEDIO  = "2F5496"
C_AZUL_CLARO  = "D9E1F2"
C_CINZA_CLARO = "F2F2F2"
C_BRANCO      = "FFFFFF"
C_VERDE       = "375623"
C_VERDE_CLARO = "E2EFDA"
C_LARANJA     = "C55A11"
C_AMARELO_CL  = "FFF2CC"
C_VERMELHO    = "C00000"
C_INPUT_BG    = "EBF3FB"
C_HEADER_FONT = C_BRANCO

TH = Font(name="Calibri", color=C_HEADER_FONT, bold=True, size=10)
TB = Font(name="Calibri", size=10)
TBB = Font(name="Calibri", bold=True, size=10)
TITLE_F = Font(name="Calibri", color=C_AZUL_ESCURO, bold=True, size=14)
SUBTIT_F = Font(name="Calibri", color=C_AZUL_MEDIO, bold=True, size=11)

def _hfill(color): return PatternFill("solid", fgColor=color)
def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center")
def _right():  return Alignment(horizontal="right",  vertical="center")

def _hdr(ws, cell, val, bg=C_AZUL_MEDIO):
    ws[cell] = val
    ws[cell].font = TH
    ws[cell].fill = _hfill(bg)
    ws[cell].alignment = _center()
    ws[cell].border = _border()

def _label(ws, cell, val):
    ws[cell].value = val
    ws[cell].font = TBB
    ws[cell].alignment = _left()
    ws[cell].border = _border("thin")

def _input_cell(ws, cell, val=None, fmt=None, bg=C_INPUT_BG):
    if val is not None:
        ws[cell].value = val
    ws[cell].fill = _hfill(bg)
    ws[cell].border = _border()
    ws[cell].alignment = _left()
    ws[cell].font = TB
    if fmt:
        ws[cell].number_format = fmt

def _result_cell(ws, cell, val=None, fmt=None):
    if val is not None:
        ws[cell].value = val
    ws[cell].fill = _hfill(C_VERDE_CLARO)
    ws[cell].border = _border()
    ws[cell].alignment = _right()
    ws[cell].font = Font(name="Calibri", bold=True, size=11, color=C_VERDE)
    if fmt:
        ws[cell].number_format = fmt


# ============================================================================
# ABA 1: CARTEIRA
# ============================================================================
def build_carteira(wb, bonds_data: list[dict]):
    ws = wb.create_sheet("Carteira")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 7
    ws.column_dimensions["K"].width = 18

    # Título
    ws.merge_cells("A1:K1")
    ws["A1"] = f"CARTEIRA DE RENDA FIXA  —  {dt_date.today().strftime('%d/%m/%Y')}"
    ws["A1"].font = TITLE_F
    ws["A1"].alignment = _center()
    ws["A1"].fill = _hfill(C_AZUL_CLARO)
    ws.row_dimensions[1].height = 28

    # Cabeçalhos
    headers = ["Ticker","Tipo","Indexador","Emissor","Vencimento",
               "VNA","Taxa Cont. (%)","PU (hoje)","Duration (anos)","Status","Sync"]
    for i, h in enumerate(headers, 1):
        _hdr(ws, f"{get_column_letter(i)}2", h)
    ws.row_dimensions[2].height = 22

    today = dt_date.today().isoformat()

    for r, b in enumerate(bonds_data, 3):
        ticker = b["ticker"]
        row_bg = C_CINZA_CLARO if r % 2 == 0 else C_BRANCO

        def cell(col, val, fmt=None, bold=False, color=None):
            ref = f"{get_column_letter(col)}{r}"
            ws[ref].value = val
            ws[ref].fill = _hfill(row_bg)
            ws[ref].border = _border("hair")
            ws[ref].alignment = _center()
            ws[ref].font = Font(name="Calibri", bold=bold, size=10,
                                color=color or "000000")
            if fmt:
                ws[ref].number_format = fmt

        cell(1, ticker, bold=True, color=C_AZUL_MEDIO)
        cell(2, b.get("tipo", ""))
        cell(3, b.get("method", ""))
        cell(4, (b.get("issuer") or "")[:40])
        cell(5, b.get("expiredate", ""), fmt="YYYY-MM-DD")
        cell(6, b.get("vna"), fmt="#,##0.0000")
        cell(7, b.get("yield_contract"), fmt="0.0000%")
        ws[f"G{r}"].number_format = "0.0000"

        # PU e duration pré-calculados
        pu_val  = b.get("pu_hoje")
        dur_val = b.get("duration_hoje")
        cell(8, pu_val,  fmt="#,##0.0000", bold=True)
        cell(9, dur_val, fmt="0.0000", bold=True)

        status = b.get("status","")
        color_status = C_VERDE if status == "A" else C_VERMELHO
        cell(10, status, bold=True, color=color_status)
        cell(11, b.get("last_sync", ""))

    # Freeze panes
    ws.freeze_panes = "A3"

    # Formatação condicional para duration
    last_row = 2 + len(bonds_data)
    ws.add_table = None  # evita conflito

    ws.row_dimensions[1].height = 30


# ============================================================================
# ABA 2: CALCULADORA
# ============================================================================
def build_calculadora(wb, tickers: list[str]):
    ws = wb.create_sheet("Calculadora")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    today_str = dt_date.today().isoformat()

    # Título
    ws.merge_cells("B1:F1")
    ws["B1"] = "CALCULADORA DE RENDA FIXA"
    ws["B1"].font = TITLE_F
    ws["B1"].alignment = _center()
    ws["B1"].fill = _hfill(C_AZUL_CLARO)
    ws.row_dimensions[1].height = 32

    # Subtítulo
    ws.merge_cells("B2:F2")
    ws["B2"] = "Preencha os inputs (fundo azul) e os resultados serão calculados automaticamente."
    ws["B2"].font = Font(name="Calibri", italic=True, size=9, color="595959")
    ws["B2"].alignment = _left()

    ws.row_dimensions[3].height = 8

    # ── Bloco de INPUTS ──────────────────────────────────────────────────────
    ws.merge_cells("B4:B4")

    def section_title(row, txt):
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"] = txt
        ws[f"B{row}"].font = Font(name="Calibri", bold=True, size=11, color=C_BRANCO)
        ws[f"B{row}"].fill = _hfill(C_AZUL_MEDIO)
        ws[f"B{row}"].alignment = _left()
        ws.row_dimensions[row].height = 22

    section_title(4, "  INPUTS")

    rows_input = [
        (5,  "Ticker",              None,          "ticker"),
        (6,  "Data de Cálculo",     today_str,     "date"),
        (7,  "Yield entrada (% a.a.)", None,       "yield"),
        (8,  "PU entrada",          None,          "pu_in"),
    ]

    for row, label, default, key in rows_input:
        _label(ws, f"B{row}", label)
        _input_cell(ws, f"C{row}", default)
        ws.row_dimensions[row].height = 22

    # Validação de dados: dropdown de tickers
    if tickers:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(tickers[:50])}"',
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(dv)
        dv.add(ws["C5"])

    # ── Bloco de RESULTADOS ──────────────────────────────────────────────────
    ws.row_dimensions[9].height = 8

    section_title(10, "  RESULTADOS  (Taxa → PU  e  PU → Taxa)")

    result_rows = [
        (11, "PU  (dado o yield da linha 7)",       "#,##0.000000"),
        (12, "Yield implícita no PU (linha 8)",      "0.000000"),
        (13, "Duration (anos)  @ yield linha 7",     "0.0000"),
        (14, "VNA atual",                             "#,##0.000000"),
        (15, "N.º de fluxos",                         "0"),
        (16, "Fonte dos dados",                       "@"),
    ]

    for row, label, fmt in result_rows:
        _label(ws, f"B{row}", label)
        _result_cell(ws, f"C{row}", fmt=fmt)
        ws.row_dimensions[row].height = 22

    # ── Fórmulas xlwings UDF ─────────────────────────────────────────────────
    ws["C11"] = '=IF(C5="","",RF_PU(C5,C7,C6))'
    ws["C12"] = '=IF(C5="","",RF_TAXA(C5,C8,C6))'
    ws["C13"] = '=IF(C5="","",RF_DURATION(C5,C7,C6))'
    ws["C14"] = '=IF(C5="","",RF_INFO(C5,"vna"))'
    ws["C15"] = '=IF(C5="","",RF_INFO(C5,"n_flows"))'
    ws["C16"] = '=IF(C5="","",RF_INFO(C5,"source"))'

    ws.row_dimensions[17].height = 8

    # ── Nota sobre xlwings ────────────────────────────────────────────────────
    section_title(18, "  COMO ATIVAR OS CÁLCULOS AUTOMÁTICOS")

    notas = [
        "1. Abra o terminal e execute:  xlwings addin install",
        "2. Reinicie o Excel. Na aba xlwings, defina o Interpreter como o caminho do Python.",
        "3. Clique em 'Run main' ou pressione Ctrl+Alt+F9 para recalcular.",
        "4. As fórmulas RF_PU / RF_TAXA / RF_DURATION serão calculadas localmente (sem API).",
        "   Código Python: sistema/excel/addin.py",
    ]
    for i, nota in enumerate(notas, 19):
        ws.merge_cells(f"B{i}:F{i}")
        ws[f"B{i}"] = nota
        ws[f"B{i}"].font = Font(name="Calibri", italic=True, size=9, color="404040")
        ws[f"B{i}"].alignment = _left()
        ws.row_dimensions[i].height = 18


# ============================================================================
# ABA 3: CADASTRAR NOVO ATIVO
# ============================================================================
def build_cadastrar(wb):
    ws = wb.create_sheet("Cadastrar")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 28

    def section_title(row, txt, color=C_AZUL_MEDIO):
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = txt
        ws[f"B{row}"].font = Font(name="Calibri", bold=True, size=11, color=C_BRANCO)
        ws[f"B{row}"].fill = _hfill(color)
        ws[f"B{row}"].alignment = _left()
        ws.row_dimensions[row].height = 22

    # Título
    ws.merge_cells("B1:E1")
    ws["B1"] = "CADASTRAR NOVO ATIVO"
    ws["B1"].font = TITLE_F
    ws["B1"].alignment = _center()
    ws["B1"].fill = _hfill(C_AZUL_CLARO)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("B2:E2")
    ws["B2"] = "Digite o ticker, execute o comando Python abaixo e os dados serão preenchidos automaticamente."
    ws["B2"].font = Font(name="Calibri", italic=True, size=9, color="595959")

    ws.row_dimensions[3].height = 8
    section_title(4, "  PASSO 1 — Digite o Ticker")

    _label(ws, "B5", "Ticker do ativo")
    _input_cell(ws, "C5", "Ex: XPTO11")
    ws.row_dimensions[5].height = 24

    # Comando Python a executar
    ws.row_dimensions[6].height = 8
    section_title(7, "  PASSO 2 — Buscar na B3 (execute no terminal)")

    cmd_rows = [
        (8,  "Comando Python:", "python main.py add {ticker}"),
        (9,  "Ou use a UDF:  ", "=RF_FETCH(C5)  [retorna dados em E9:E18]"),
        (10, "Ou botão:      ", "Se xlwings instalado: aba xlwings > Run main > buscar_b3"),
    ]
    for row, lbl, cmd in cmd_rows:
        _label(ws, f"B{row}", lbl)
        ws[f"C{row}"] = cmd
        ws[f"C{row}"].font = Font(name="Courier New", size=9, color=C_AZUL_MEDIO)
        ws[f"C{row}"].fill = _hfill("F5F5F5")
        ws[f"C{row}"].border = _border("hair")
        ws.row_dimensions[row].height = 20

    # Bloco de dados auto-preenchidos
    ws.row_dimensions[11].height = 8
    section_title(12, "  DADOS DO ATIVO (preenchidos após buscar na B3)")

    fields = [
        ("Tipo",              "DEB / CRI / CRA"),
        ("Indexador",         "IPCA / CDI / PRE"),
        ("Emissor",           ""),
        ("Data de Emissão",   "YYYY-MM-DD"),
        ("Data de Vencimento","YYYY-MM-DD"),
        ("VNA / VNE",         ""),
        ("Taxa do Contrato (%)", ""),
        ("Dia de Aniversário",""),
        ("Nº de Fluxos",      ""),
        ("Status",            "A = Ativo"),
    ]
    for i, (label, placeholder) in enumerate(fields, 13):
        _label(ws, f"B{i}", label)
        _input_cell(ws, f"C{i}", placeholder if placeholder else None, bg="FFFBE6")
        ws.row_dimensions[i].height = 20

    # UDF RF_FETCH popula automaticamente
    for i, _ in enumerate(fields, 13):
        ws[f"E{i}"] = f'=IFERROR(RF_INFO(C5,B{i}),"")'
        ws[f"E{i}"].font = Font(name="Calibri", size=9, color="595959", italic=True)
        ws[f"E{i}"].fill = _hfill(C_CINZA_CLARO)
        ws[f"E{i}"].border = _border("hair")

    # Passo 3: Registrar
    last = 13 + len(fields)
    ws.row_dimensions[last].height = 8
    section_title(last + 1, "  PASSO 3 — Registrar")

    ws.merge_cells(f"B{last+2}:E{last+2}")
    cmd_reg = "python main.py add {ticker}    # Registra e baixa fluxos automaticamente"
    ws[f"B{last+2}"] = cmd_reg
    ws[f"B{last+2}"].font = Font(name="Courier New", size=10, color=C_AZUL_MEDIO, bold=True)
    ws[f"B{last+2}"].fill = _hfill(C_AMARELO_CL)
    ws[f"B{last+2}"].border = _border()
    ws.row_dimensions[last + 2].height = 26


# ============================================================================
# ABA 4: DADOS HISTÓRICOS
# ============================================================================
def build_dados(wb):
    from src.db import get_conn

    ws = wb.create_sheet("Dados")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 12

    ws.merge_cells("A1:E1")
    ws["A1"] = "DADOS HISTÓRICOS"
    ws["A1"].font = TITLE_F
    ws["A1"].fill = _hfill(C_AZUL_CLARO)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    # CDI
    _hdr(ws, "A2", "Data CDI")
    _hdr(ws, "B2", "CDI (% a.d.)")
    ws.row_dimensions[2].height = 20

    with get_conn() as conn:
        cdis = conn.execute(
            "SELECT date, rate_annual FROM cdi_daily ORDER BY date DESC LIMIT 500"
        ).fetchall()
    for i, row in enumerate(cdis, 3):
        ws[f"A{i}"] = row["date"]
        ws[f"B{i}"] = float(row["rate_annual"])
        ws[f"B{i}"].number_format = "0.0000"
        fill = _hfill(C_CINZA_CLARO if i % 2 == 0 else C_BRANCO)
        ws[f"A{i}"].fill = fill
        ws[f"B{i}"].fill = fill
        ws[f"A{i}"].font = TB
        ws[f"B{i}"].font = TB

    # IPCA
    _hdr(ws, "D2", "Data IPCA")
    _hdr(ws, "E2", "IPCA (% a.m.)")
    with get_conn() as conn:
        ipcas = conn.execute(
            "SELECT date, rate_pct FROM ipca_monthly ORDER BY date DESC LIMIT 100"
        ).fetchall()
    for i, row in enumerate(ipcas, 3):
        ws[f"D{i}"] = row["date"]
        ws[f"E{i}"] = float(row["rate_pct"])
        ws[f"E{i}"].number_format = "0.0000"
        fill = _hfill(C_CINZA_CLARO if i % 2 == 0 else C_BRANCO)
        ws[f"D{i}"].fill = fill
        ws[f"E{i}"].fill = fill
        ws[f"D{i}"].font = TB
        ws[f"E{i}"].font = TB

    ws.freeze_panes = "A3"


# ============================================================================
# ENTRY POINT
# ============================================================================
def generate(output_path: str | None = None):
    from src.db import get_conn
    from src.calc import calc_pu

    path = Path(output_path) if output_path else OUTPUT_PATH

    # Coleta dados dos bonds
    with get_conn() as conn:
        bonds_raw = conn.execute(
            "SELECT * FROM bonds ORDER BY tipo, ticker"
        ).fetchall()

    bonds_data = []
    tickers = []
    today = dt_date.today().isoformat()

    for b in bonds_raw:
        info = dict(b)
        ticker = info["ticker"]
        tickers.append(ticker)
        y0 = info.get("yield_contract") or 0.0
        try:
            r = calc_pu(ticker, y0, today, fetch_if_missing=False)
            info["pu_hoje"] = r["PU"]
            info["duration_hoje"] = r["duration"]
        except Exception:
            info["pu_hoje"] = None
            info["duration_hoje"] = None
        bonds_data.append(info)

    # Cria workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove sheet padrão

    build_carteira(wb, bonds_data)
    build_calculadora(wb, tickers)
    build_cadastrar(wb)
    build_dados(wb)

    # Ativa a aba Carteira ao abrir
    wb.active = wb["Carteira"]
    wb.save(path)
    print(f"Dashboard salvo: {path}")
    return str(path)


if __name__ == "__main__":
    generate()
