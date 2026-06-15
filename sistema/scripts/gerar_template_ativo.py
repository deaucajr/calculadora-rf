"""
Gera o arquivo CADASTRO_ATIVO.xlsx — template para cadastro manual de ativos.

Uso:
  python scripts/gerar_template_ativo.py                  -> cria na pasta atual
  python scripts/gerar_template_ativo.py VALE23           -> pre-preenche ticker
  python scripts/gerar_template_ativo.py --output pasta/  -> salva em outra pasta

Depois de preencher, rode:
  python scripts/importar_planilha.py CADASTRO_ATIVO.xlsx
"""
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERRO: pip install openpyxl")
    sys.exit(1)

AZUL   = "FF1F4E79"
CINZA  = "FFD9D9D9"
VERDE  = "FF375623"
AMARELO = "FFFFF2CC"
BRANCO  = "FFFFFFFF"
LARANJA = "FFED7D31"


def fill(hex8): return PatternFill("solid", fgColor=hex8)
def bold(sz=11, cor=BRANCO): return Font(bold=True, size=sz, color=cor)
def borda():
    s = Side(style="thin", color="FF999999")
    return Border(left=s, right=s, top=s, bottom=s)
def centro(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def esq():    return Alignment(horizontal="left",   vertical="center", wrap_text=True)


def sheet_identificacao(wb, ticker_padrao=""):
    ws = wb.create_sheet("Identificacao")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38

    # Titulo
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "IDENTIFICACAO DO ATIVO"
    c.font = Font(bold=True, size=13, color=BRANCO)
    c.fill = fill(AZUL); c.alignment = centro()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = "Preencha a coluna amarela. Campos com * sao obrigatorios."
    c.font = Font(italic=True, size=9, color="FF606060")
    c.alignment = esq()

    campos = [
        ("Ticker *",              ticker_padrao,  "Ex: VALE23, EGIEA6"),
        ("Tipo *",                "DEB",          "DEB | CRI | CRA"),
        ("Indexador *",           "IPCA+",        "IPCA+ | PRE | CDI+ | %CDI"),
        ("Emissor",               "",             "Nome da empresa emissora"),
        ("Data de Emissao *",     "",             "DD/MM/AAAA — data de inicio da rentabilidade"),
        ("Vencimento *",          "",             "DD/MM/AAAA — data do ultimo pagamento"),
        ("VNE (R$) *",            1000,           "Valor Nominal de Emissao (geralmente 1000)"),
        ("Tai — Taxa de Emissao *","",             "Taxa CONTRATUAL em % a.a. (ex: 6.5 p/ IPCA+6,5% | 12.0 p/ PRE 12%)"),
        ("Rating",                "",             "Opcional: AAA, AA+, AA, AA-, A+, A, A-..."),
        ("Garantia",              "",             "Opcional: Quirografaria | Real | Fidejussoria"),
        ("Codigo CETIP/ISIN",     "",             "Opcional"),
    ]

    for i, (label, valor, dica) in enumerate(campos, 3):
        obrig = "*" in label
        r = ws.row_dimensions[i]; r.height = 20

        # coluna A — label
        ca = ws.cell(row=i, column=1, value=label)
        ca.font = Font(bold=obrig, size=10, color="FF1F4E79" if obrig else "FF404040")
        ca.fill = fill(CINZA); ca.border = borda(); ca.alignment = esq()

        # coluna B — valor (amarelo = preencher)
        cb = ws.cell(row=i, column=2, value=valor if valor != "" else None)
        cb.fill = fill(AMARELO); cb.border = borda(); cb.alignment = esq()
        cb.font = Font(size=10)
        if "Data" in label:
            cb.number_format = "DD/MM/YYYY"

        # dica como comentario (openpyxl nao suporta tooltip nativo, usa coluna C)
        cd = ws.cell(row=i, column=3, value=f"← {dica}")
        cd.font = Font(italic=True, size=8, color="FF808080")
        cd.alignment = esq()

    ws.column_dimensions["C"].width = 45

    # Validacao Indexador
    dv = DataValidation(type="list", formula1='"IPCA+,PRE,CDI+,%CDI"', allow_blank=False)
    dv.sqref = "B5"
    ws.add_data_validation(dv)

    return ws


def sheet_fluxo(wb):
    ws = wb.create_sheet("Fluxo_de_Caixa")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Titulo
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "FLUXO DE CAIXA — Cole aqui a tabela de pagamentos (Ctrl+V na linha 3)"
    c.font = Font(bold=True, size=12, color=BRANCO)
    c.fill = fill(VERDE); c.alignment = centro()
    ws.row_dimensions[1].height = 24

    # Instrucao rapida
    ws.merge_cells("A2:E2")
    inst = ws["A2"]
    inst.value = (
        "Data: DD/MM/AAAA   |   "
        "Paga Juros: S ou N   |   "
        "% Amortizacao: soma deve ser 100 (amortizante) ou deixe 0 em tudo e 100 no vencimento (bullet)   |   "
        "Tai: taxa CONTRATUAL do ativo em % a.a. — deixe em BRANCO se for igual em todos os periodos (usa o valor da aba Identificacao)"
    )
    inst.font = Font(italic=True, size=8, color="FF404040")
    inst.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    # Cabecalho das colunas
    # Coluna D (Tai) fica cinza para indicar que e opcional
    hdrs    = ["Data", "Paga Juros? (S/N)", "% Amortizacao", "Tai — Taxa Contratual (% a.a.)"]
    larguras = [16,     22,                  20,               30]
    fills_hdr = [AZUL, AZUL, AZUL, "FF4A4A4A"]  # coluna Tai mais escura = opcional
    for col, (h, w, fh) in enumerate(zip(hdrs, larguras, fills_hdr), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = bold(10, BRANCO); c.fill = fill(fh)
        c.border = borda(); c.alignment = centro()
        ws.column_dimensions[get_column_letter(col)].width = w

    # Legenda na coluna E
    ws.column_dimensions["E"].width = 48
    notas_col = [
        (3, "← DEIXAR EM BRANCO na maioria dos casos"),
        (4, "← Em branco = usa Tai da aba Identificacao"),
        (8, "← Preencha so em bonds com step-up (taxa muda por periodo)"),
    ]
    for row_n, txt in notas_col:
        c = ws.cell(row=row_n, column=5, value=txt)
        c.font = Font(italic=True, size=8, color="FF808080")

    ws.row_dimensions[3].height = 22

    # Linhas de exemplo (cinza claro, para o usuario ver o formato e apagar)
    exemplos = [
        ("15/01/2027", "S", 0,   None),   # Tai em branco = usa da identificacao
        ("15/07/2027", "S", 0,   None),
        ("15/01/2028", "S", 0,   None),
        ("15/07/2028", "S", 25,  None),
        ("15/01/2029", "S", 25,  None),
        ("15/07/2029", "S", 25,  None),
        ("15/01/2030", "S", 25,  None),
        ("15/07/2030", "S", 25,  None),
    ]
    for i, (dt, j, amort, taxa) in enumerate(exemplos, 4):
        fmts = ["DD/MM/YYYY", "@", "0.00", "0.00##"]
        vals = [dt, j, amort, taxa]
        for col, (v, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = Font(italic=True, size=10, color="FF808080")
            c.fill = fill("FFF5F5F5")
            c.border = borda(); c.alignment = centro()
            c.number_format = fmt

    # Nota sobre os exemplos
    ws.merge_cells("A12:D12")
    nota = ws["A12"]
    nota.value = "^ EXEMPLOS — apague essas linhas e cole os dados reais a partir da linha 4 ^"
    nota.font = Font(italic=True, size=8, color="FFED7D31", bold=True)
    nota.alignment = centro()

    # Validacao coluna B (S/N)
    dv_sn = DataValidation(type="list", formula1='"S,N"', allow_blank=True)
    dv_sn.sqref = "B4:B500"
    ws.add_data_validation(dv_sn)

    return ws


def sheet_ajuda(wb):
    ws = wb.create_sheet("Ajuda")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 90

    ws.merge_cells("A1:A1")
    ws["A1"].value = "COMO USAR O CADASTRO_ATIVO.xlsx"
    ws["A1"].font = Font(bold=True, size=14, color=BRANCO)
    ws["A1"].fill = fill(AZUL); ws["A1"].alignment = centro()
    ws.row_dimensions[1].height = 28

    linhas = [
        "",
        "PASSO 1 — Aba 'Identificacao'",
        "  Preencha os campos amarelos. Os campos com * sao obrigatorios.",
        "  Indexador: use exatamente IPCA+ | PRE | CDI+ | %CDI",
        "  Taxa de Emissao: a taxa CONTRATUAL do ativo (ex: 6.5 para IPCA+6,5% a.a.)",
        "  VNE: valor nominal de emissao — quase sempre R$ 1.000,00",
        "",
        "PASSO 2 — Aba 'Fluxo_de_Caixa'",
        "  Apague as linhas de exemplo (linhas 4 a 11).",
        "  Cole a tabela de pagamentos a partir da linha 4.",
        "  Voce pode copiar do Bloomberg, de uma planilha propria ou do prospecto.",
        "",
        "  COLUNAS:",
        "    Data           : data de cada pagamento no formato DD/MM/AAAA",
        "    Paga Juros     : S se paga cupom nessa data, N se so amortiza",
        "    % Amortizacao  : quanto do principal e pago (0 a 100).",
        "                     Bullet: todos 0 exceto o vencimento que tem 100.",
        "                     Amortizante: ex 25, 25, 25, 25 (soma = 100).",
        "    Tai (% a.a.)   : Taxa CONTRATUAL do ativo — define o tamanho do cupom.",
        "                     Para IPCA+6,5%: coloque 6.5.  Para PRE 12%: coloque 12.",
        "                     DEIXE EM BRANCO na maioria dos casos: o script usa o valor",
        "                     de 'Taxa de Emissao' preenchido na aba Identificacao.",
        "                     Preencha so em bonds com step-up (taxa muda por periodo).",
        "",
        "  IMPORTANTE — TAI x TAXA DE MERCADO:",
        "    Tai    = taxa CONTRATUAL (define o fluxo de caixa — vai para o CSV)",
        "    Taxa de mercado = usada depois no Excel: =RF_PU(ticker; TAXA_MERCADO; data)",
        "    Sao coisas diferentes! O Tai e fixo no prospecto; a de mercado oscila.",
        "",
        "  EXEMPLOS:",
        "  Bullet IPCA+6,5%:      Data      | S | 0  | (branco)",
        "                          Data      | S | 0  | (branco)",
        "                          Vencimento| S | 100| (branco)",
        "",
        "  Amortizante PRE 12%:   Data      | S | 25 | 12.0",
        "                          Data      | S | 25 | 12.0",
        "                          Data      | S | 25 | 12.0",
        "                          Vencimento| S | 25 | 12.0",
        "",
        "PASSO 3 — Importar",
        "  Com o arquivo salvo, abra o terminal na pasta 'sistema' e rode:",
        "    python scripts/importar_planilha.py CADASTRO_ATIVO.xlsx",
        "",
        "  O script cria automaticamente o arquivo CSV em fluxos_manual/",
        "  e o ativo fica disponivel nas funcoes RF_* sem precisar reabrir o Excel.",
        "",
        "DICAS",
        "  - O ticker deve ser unico (ex: nao use o mesmo ticker de um ativo da B3).",
        "  - Para CDI+ e %CDI, use a importacao automatica da B3 (importar_fluxos.py).",
        "    O cadastro manual de CDI nao e suportado pois depende da curva diaria.",
        "  - Se errar, basta salvar o Excel corrigido e rodar o importar_planilha.py de novo.",
        "  - Para atualizar o VNA (IPCA): rode rotina_diaria.py no dia a dia.",
    ]
    for i, txt in enumerate(linhas, 2):
        c = ws.cell(row=i, column=1, value=txt)
        if txt and not txt.startswith(" ") and not txt.startswith("  "):
            c.font = Font(bold=True, size=10, color="FF1F4E79")
        else:
            c.font = Font(size=10, color="FF303030")
        c.alignment = esq()

    return ws


def gerar(ticker="", output_dir=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_identificacao(wb, ticker)
    sheet_fluxo(wb)
    sheet_ajuda(wb)

    if output_dir is None:
        out = Path.cwd() / "CADASTRO_ATIVO.xlsx"
    else:
        out = Path(output_dir) / "CADASTRO_ATIVO.xlsx"

    wb.save(str(out))
    print(f"Template criado: {out}")
    print("Preencha as abas 'Identificacao' e 'Fluxo_de_Caixa', depois rode:")
    print("  python scripts/importar_planilha.py CADASTRO_ATIVO.xlsx")
    return out


if __name__ == "__main__":
    args = sys.argv[1:]
    ticker = ""
    output = None
    for i, a in enumerate(args):
        if a == "--output" and i + 1 < len(args):
            output = args[i + 1]
        elif not a.startswith("--"):
            ticker = a.upper()
    gerar(ticker, output)
