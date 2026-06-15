"""
Gera planilha Excel de analise de swap entre indexadores de renda fixa brasileira.

Cadeia de equivalencia:
  %CDI  <-A->  CDI+spread  <-B->  PRE  <-C->  IPCA+spread

Formulas (Gemini 2.5 Pro / ANBIMA/B3 confirmado):
  A: spread_cdi = [(1+CDI/100)^(pct/100-1) - 1]*100  (independente do prazo)
     pct_cdi    = [1 + ln(1+spread/100)/ln(1+CDI/100)]*100
  B: CDI_leg_nominal = (1+CDI/100)*(1+spread_cdi/100)  [multiplicativo]
  C: spread_ipca = [CDI_leg_nominal/(1+infla/100) - 1]*100  (Fisher)
     ou inverso: spread_cdi = [(1+spread_ipca/100)*(1+infla/100)/(1+CDI/100)... - 1]*100

Breakeven:
  IPCA+ vs PRE: inflacao_implicita = (1+taxaPRE/100)/(1+taxaIPCA/100) - 1  (Fisher)
  %CDI  vs PRE: CDI_breakeven = (1+taxaPRE/100)^(100/pct) - 1

NII por periodo:
  CDI leg : nocional * [(1+CDI/100)*(1+spread/100)]^(du/252) - nocional
  IPCA leg: nocional * (1+infla_periodo) * (1+spread_ipca/100)^(du/252) - nocional
  Resultado (net): NII_IPCA - NII_CDI (positivo = recebe mais no lado IPCA)

Uso:
  python scripts/gerar_planilha_swap.py
  python scripts/gerar_planilha_swap.py --output meu_swap.xlsx
  python scripts/gerar_planilha_swap.py --cdi 13.5 --ipca 5.2 --spread-cdi 2.0 --nocional 10000000
"""
import sys
import math
import datetime
from pathlib import Path

# Tenta importar openpyxl; instrucoes se ausente
try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers as xl_numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    print("ERRO: openpyxl nao instalado. Execute: pip install openpyxl")
    sys.exit(1)


# ── Formulas financeiras ──────────────────────────────────────────────────────

def pct_para_cdi(pct: float, cdi: float) -> float:
    """
    %CDI -> spread CDI+ equivalente (% a.a.)
    spread = [(1+CDI/100)^(pct/100-1) - 1]*100
    Independe do prazo.
    Ex: pct=98, CDI=13.5% -> spread ≈ -0.18%
    """
    return ((1 + cdi / 100) ** (pct / 100 - 1) - 1) * 100


def cdi_para_pct(spread: float, cdi: float) -> float:
    """
    Spread CDI+ -> %CDI equivalente
    pct = [1 + ln(1+spread/100)/ln(1+CDI/100)] * 100
    """
    if cdi <= 0:
        return float("nan")
    return (1 + math.log(1 + spread / 100) / math.log(1 + cdi / 100)) * 100


def cdi_para_ipca(spread_cdi: float, cdi: float, infla: float) -> float:
    """
    CDI+spread -> spread IPCA+ equivalente (Fisher multiplicativo, % a.a.)
    leg_cdi_nominal = (1+CDI/100)*(1+spread/100)
    spread_ipca = [leg_cdi_nominal/(1+infla/100) - 1]*100
    """
    return ((1 + cdi / 100) * (1 + spread_cdi / 100) / (1 + infla / 100) - 1) * 100


def ipca_para_cdi(spread_ipca: float, cdi: float, infla: float) -> float:
    """IPCA+spread -> spread CDI+ equivalente"""
    return ((1 + spread_ipca / 100) * (1 + infla / 100) / (1 + cdi / 100) - 1) * 100


def pct_para_ipca(pct: float, cdi: float, infla: float) -> float:
    """
    %CDI -> spread IPCA+ equivalente (encadeia A+C)
    (1+CDI/100)^(pct/100) = (1+spread_ipca/100)*(1+infla/100)
    spread_ipca = [(1+CDI/100)^(pct/100)/(1+infla/100) - 1]*100
    """
    return ((1 + cdi / 100) ** (pct / 100) / (1 + infla / 100) - 1) * 100


def ipca_para_pct(spread_ipca: float, cdi: float, infla: float) -> float:
    """IPCA+spread -> %CDI equivalente"""
    if cdi <= 0:
        return float("nan")
    nominal = (1 + spread_ipca / 100) * (1 + infla / 100)
    return math.log(nominal) / math.log(1 + cdi / 100) * 100


def infla_implicita(taxa_pre: float, taxa_ipca: float) -> float:
    """Breakeven PRE vs IPCA+: inflacao implicita (Fisher)."""
    return ((1 + taxa_pre / 100) / (1 + taxa_ipca / 100) - 1) * 100


def cdi_breakeven_pct(taxa_pre: float, pct_cdi: float) -> float:
    """Breakeven %CDI vs PRE: CDI medio que iguala os retornos."""
    return ((1 + taxa_pre / 100) ** (100 / pct_cdi) - 1) * 100


def nii_periodo(nocional: float, taxa_aa: float, du: int) -> float:
    """
    NII (Net Interest Income) de um periodo em R$.
    nocional * [(1+taxa/100)^(du/252) - 1]
    """
    return nocional * ((1 + taxa_aa / 100) ** (du / 252) - 1)


def gerar_datas_semestrais(data_ini: datetime.date, data_fim: datetime.date) -> list[datetime.date]:
    """Datas de pagamento semestrais (cada 6 meses) ate o vencimento."""
    datas = []
    d = data_ini
    while d < data_fim:
        mes = d.month + 6
        ano = d.year + (mes - 1) // 12
        mes = (mes - 1) % 12 + 1
        d = datetime.date(ano, mes, min(d.day, 28))
        if d <= data_fim:
            datas.append(d)
    if not datas or datas[-1] != data_fim:
        datas.append(data_fim)
    return datas


def du_util_aprox(d0: datetime.date, d1: datetime.date) -> int:
    """Dias uteis aproximados (sem feriados — para NII ilustrativo)."""
    dc = (d1 - d0).days
    semanas = dc // 7
    resto = dc % 7
    dow = d0.weekday()
    dias_uteis = semanas * 5
    for i in range(resto):
        if (dow + i) % 7 < 5:
            dias_uteis += 1
    return max(dias_uteis, 1)


# ── Estilos Excel ─────────────────────────────────────────────────────────────

def cor(hex_str: str) -> str:
    h = hex_str.lstrip("#").upper()
    return "FF" + h if len(h) == 6 else h


AZUL_ESCURO = cor("#1F4E79")
AZUL_MEDIO  = cor("#2E75B6")
AZUL_CLARO  = cor("#D6E4F0")
VERDE       = cor("#375623")
VERDE_CLARO = cor("#E2EFDA")
AMARELO     = cor("#FFF2CC")
LARANJA     = cor("#F4B942")
CINZA       = cor("#D9D9D9")
BRANCO      = cor("#FFFFFF")

def hdr_fill(hex_c: str): return PatternFill("solid", fgColor=hex_c)
def bold(size: int = 11): return Font(bold=True, size=size)
def titulo_font(): return Font(bold=True, size=12, color=BRANCO)
def centro(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def direita(): return Alignment(horizontal="right")

def borda_fina():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def aplicar_cabecalho(ws, linha: int, textos: list, fill_hex: str = AZUL_MEDIO):
    fill = hdr_fill(fill_hex)
    f = Font(bold=True, color=BRANCO, size=10)
    for col, txt in enumerate(textos, 1):
        c = ws.cell(row=linha, column=col, value=txt)
        c.fill = fill; c.font = f; c.alignment = centro(); c.border = borda_fina()


def set_larguras(ws, larguras: dict):
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w


# ── Planilha 1: Equivalencias de Indexador ────────────────────────────────────

def sheet_equivalencias(wb, cdi: float, infla: float):
    ws = wb.create_sheet("Equivalencias")
    ws.sheet_view.showGridLines = False

    # Titulo
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "TABELA DE EQUIVALÊNCIA ENTRE INDEXADORES"
    c.font = Font(bold=True, size=14, color=BRANCO)
    c.fill = hdr_fill(AZUL_ESCURO); c.alignment = centro()
    ws.row_dimensions[1].height = 28

    # Parametros
    ws["A3"] = "CDI Projetado (% a.a.):"; ws["A3"].font = bold()
    ws["B3"] = cdi; ws["B3"].number_format = '0.00"%"'; ws["B3"].fill = hdr_fill(AMARELO)
    ws["A4"] = "IPCA Projetado (% a.a.):"; ws["A4"].font = bold()
    ws["B4"] = infla; ws["B4"].number_format = '0.00"%"'; ws["B4"].fill = hdr_fill(AMARELO)
    ws["D3"] = "*(altere os valores amarelos para recalcular)*"
    ws["D3"].font = Font(italic=True, size=9, color="FF808080")

    # Tabela A: %CDI -> CDI+ e IPCA+
    ws["A6"] = "BLOCO A — %CDI → CDI+spread → IPCA+spread"
    ws["A6"].font = Font(bold=True, size=11)
    aplicar_cabecalho(ws, 7, ["%CDI", "Spread CDI+ (% a.a.)", "Spread IPCA+ (% a.a.)", "Inflação Implícita"])
    pcts = [85, 90, 95, 97, 98, 99, 100, 101, 102, 105, 110, 115, 120]
    for i, pct in enumerate(pcts, 8):
        sc = pct_para_cdi(pct, cdi)
        si = pct_para_ipca(pct, cdi, infla)
        inf_impl = infla_implicita(cdi + sc, si) if abs(si) < 50 else 0
        fill = hdr_fill(VERDE_CLARO) if pct < 100 else hdr_fill(AZUL_CLARO) if pct == 100 else hdr_fill(AMARELO)
        for col, val, fmt in [
            (1, pct, '0.0"%"'),
            (2, sc, '+0.000"%";-0.000"%"'),
            (3, si, '0.0000"%"'),
            (4, inf_impl, '0.00"%"'),
        ]:
            c = ws.cell(row=i, column=col, value=val)
            c.number_format = fmt; c.fill = fill; c.border = borda_fina()
            c.alignment = centro()

    # Tabela B: CDI+ -> IPCA+ (varios spreads)
    linha_b = 8 + len(pcts) + 2
    ws.cell(row=linha_b, column=1, value="BLOCO B — CDI+spread → IPCA+spread equivalente")
    ws.cell(row=linha_b, column=1).font = Font(bold=True, size=11)
    aplicar_cabecalho(ws, linha_b + 1,
        ["Spread CDI+ (% a.a.)", "Nominal CDI leg (% a.a.)", "Spread IPCA+ (% a.a.)", "Inflação Implícita"])
    spreads_cdi = [-0.5, 0, 0.25, 0.5, 0.75, 1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0]
    for i, sc in enumerate(spreads_cdi, linha_b + 2):
        nominal = ((1 + cdi / 100) * (1 + sc / 100) - 1) * 100
        si = cdi_para_ipca(sc, cdi, infla)
        inf_i = infla_implicita(nominal, si)
        fill = hdr_fill(AZUL_CLARO)
        for col, val, fmt in [
            (1, sc, '+0.000"%";-0.000"%"'),
            (2, nominal, '0.0000"%"'),
            (3, si, '0.0000"%"'),
            (4, inf_i, '0.00"%"'),
        ]:
            c = ws.cell(row=i, column=col, value=val)
            c.number_format = fmt; c.fill = fill; c.border = borda_fina()
            c.alignment = centro()

    # Larguras
    set_larguras(ws, {"A": 28, "B": 22, "C": 22, "D": 22, "E": 16, "F": 16, "G": 16, "H": 16})
    return ws


# ── Planilha 2: Swap CDI+ <-> IPCA+ ──────────────────────────────────────────

def sheet_swap_cdi_ipca(wb, cdi: float, infla: float, spread_cdi: float,
                        nocional: float, data_ini: datetime.date, data_fim: datetime.date):
    ws = wb.create_sheet("Swap CDI+ ↔ IPCA+")
    ws.sheet_view.showGridLines = False

    spread_ipca = cdi_para_ipca(spread_cdi, cdi, infla)
    nominal_cdi = ((1 + cdi / 100) * (1 + spread_cdi / 100) - 1) * 100

    # Titulo
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"SWAP: CDI + {spread_cdi:.2f}% a.a.  ↔  IPCA + {spread_ipca:.4f}% a.a."
    c.font = Font(bold=True, size=13, color=BRANCO)
    c.fill = hdr_fill(AZUL_ESCURO); c.alignment = centro()
    ws.row_dimensions[1].height = 26

    # Parametros
    params = [
        ("Nocional (R$)",       nocional,    '#,##0.00'),
        ("Data Início",         data_ini,    'DD/MM/YYYY'),
        ("Data Fim",            data_fim,    'DD/MM/YYYY'),
        ("CDI Projetado",       cdi,         '0.00"%"'),
        ("IPCA Projetado",      infla,       '0.00"%"'),
        ("Spread CDI+ (input)", spread_cdi,  '+0.00"%";-0.00"%"'),
        ("Spread IPCA+ (equiv)",spread_ipca, '0.0000"%"'),
        ("Taxa nominal CDI leg", nominal_cdi,'0.0000"%"'),
    ]
    for i, (label, val, fmt) in enumerate(params, 3):
        ws.cell(row=i, column=1, value=label).font = bold()
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = fmt
        c.fill = hdr_fill(AMARELO if "input" in label.lower() or label.startswith("CDI Proj") or label.startswith("IPCA Proj") else AZUL_CLARO)

    # Tabela NII
    linha_nii = 3 + len(params) + 2
    ws.cell(row=linha_nii, column=1, value="PROJEÇÃO DE NII POR PERÍODO (SEMESTRAL)").font = Font(bold=True, size=11)
    hdrs = ["Período", "Data Pgto", "DU (aprox.)", "NII Leg CDI (R$)", "NII Leg IPCA (R$)",
            "Net (IPCA - CDI)", "Acumulado (R$)", "% do Nocional"]
    aplicar_cabecalho(ws, linha_nii + 1, hdrs)

    datas = gerar_datas_semestrais(data_ini, data_fim)
    acum = 0.0
    prev = data_ini
    for i, d in enumerate(datas, 1):
        du = du_util_aprox(prev, d)
        nii_cdi  = nii_periodo(nocional, nominal_cdi, du)
        nii_ipca = nii_periodo(nocional, spread_ipca + infla, du)  # aproximado: (spread+infla)
        net = nii_ipca - nii_cdi
        acum += net
        row = linha_nii + 1 + i
        fill = hdr_fill(VERDE_CLARO) if net >= 0 else hdr_fill(cor("#FFCCCC"))
        vals = [i, d, du, nii_cdi, nii_ipca, net, acum, net / nocional * 100]
        fmts = ['0', 'DD/MM/YYYY', '0', '#,##0.00', '#,##0.00', '+#,##0.00;-#,##0.00', '#,##0.00', '+0.0000"%";-0.0000"%"']
        for col, (v, f) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=row, column=col, value=v)
            c.number_format = f; c.fill = fill; c.border = borda_fina()
            c.alignment = centro() if col in (1, 3) else direita()
        prev = d

    # Total
    row_tot = linha_nii + 1 + len(datas) + 1
    ws.cell(row=row_tot, column=1, value="TOTAL").font = bold()
    ws.cell(row=row_tot, column=7, value=acum).number_format = '#,##0.00'
    ws.cell(row=row_tot, column=7).font = bold()
    ws.cell(row=row_tot, column=8, value=acum / nocional * 100).number_format = '+0.0000"%";-0.0000"%"'
    ws.cell(row=row_tot, column=8).font = bold()

    set_larguras(ws, {"A": 20, "B": 14, "C": 14, "D": 18, "E": 18, "F": 18, "G": 18, "H": 16, "I": 14})
    return ws


# ── Planilha 3: Swap %CDI <-> IPCA+ ──────────────────────────────────────────

def sheet_swap_pct_ipca(wb, cdi: float, infla: float, pct_cdi: float,
                        nocional: float, data_ini: datetime.date, data_fim: datetime.date):
    ws = wb.create_sheet("Swap %CDI ↔ IPCA+")
    ws.sheet_view.showGridLines = False

    spread_cdi_eq = pct_para_cdi(pct_cdi, cdi)
    spread_ipca_eq = pct_para_ipca(pct_cdi, cdi, infla)
    nominal_pct = ((1 + cdi / 100) ** (pct_cdi / 100) - 1) * 100

    # Titulo
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"SWAP: {pct_cdi:.1f}% do CDI  ↔  CDI + {spread_cdi_eq:.4f}%  ↔  IPCA + {spread_ipca_eq:.4f}%"
    c.font = Font(bold=True, size=13, color=BRANCO)
    c.fill = hdr_fill(VERDE); c.alignment = centro()
    ws.row_dimensions[1].height = 26

    params = [
        ("Nocional (R$)",        nocional,       '#,##0.00'),
        ("Data Início",          data_ini,       'DD/MM/YYYY'),
        ("Data Fim",             data_fim,       'DD/MM/YYYY'),
        ("CDI Projetado",        cdi,            '0.00"%"'),
        ("IPCA Projetado",       infla,          '0.00"%"'),
        ("% do CDI (input)",     pct_cdi,        '0.00"%"'),
        ("→ Spread CDI+ equiv.", spread_cdi_eq,  '+0.0000"%";-0.0000"%"'),
        ("→ Spread IPCA+ equiv.",spread_ipca_eq, '0.0000"%"'),
        ("Taxa nominal %CDI",    nominal_pct,    '0.0000"%"'),
    ]
    for i, (label, val, fmt) in enumerate(params, 3):
        ws.cell(row=i, column=1, value=label).font = bold()
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = fmt
        c.fill = hdr_fill(AMARELO if "input" in label.lower() or "Projetado" in label else VERDE_CLARO)

    # Tabela equivalencias %CDI x varios CDI
    linha_tab = 3 + len(params) + 2
    ws.cell(row=linha_tab, column=1, value=f"SENSIBILIDADE: Spread IPCA+ equivalente a {pct_cdi:.0f}% CDI por nível de CDI").font = Font(bold=True, size=11)
    aplicar_cabecalho(ws, linha_tab + 1, ["CDI Proj. (%)", "Spread CDI+", "Spread IPCA+",
                                           "Nominal %CDI (%)", "Infla Implícita (IPCA+vsNominal)"])
    cdis_cenarios = [8, 9, 10, 10.5, 11, 11.5, 12, 12.5, 13, 13.5, 14, 15, 16]
    for i, cdi_s in enumerate(cdis_cenarios, linha_tab + 2):
        sc_eq = pct_para_cdi(pct_cdi, cdi_s)
        si_eq = pct_para_ipca(pct_cdi, cdi_s, infla)
        nom   = ((1 + cdi_s / 100) ** (pct_cdi / 100) - 1) * 100
        inf_i = infla_implicita(nom, si_eq) if si_eq > -99 else 0
        fill = hdr_fill(AMARELO) if abs(cdi_s - cdi) < 0.01 else hdr_fill(VERDE_CLARO)
        for col, val, fmt in [
            (1, cdi_s, '0.0"%"'),
            (2, sc_eq, '+0.0000"%";-0.0000"%"'),
            (3, si_eq, '0.0000"%"'),
            (4, nom,   '0.0000"%"'),
            (5, inf_i, '0.00"%"'),
        ]:
            c = ws.cell(row=i, column=col, value=val)
            c.number_format = fmt; c.fill = fill; c.border = borda_fina()
            c.alignment = centro()

    set_larguras(ws, {"A": 22, "B": 18, "C": 18, "D": 20, "E": 24, "F": 16})
    return ws


# ── Planilha 4: Breakeven ─────────────────────────────────────────────────────

def sheet_breakeven(wb, cdi: float, infla: float):
    ws = wb.create_sheet("Breakeven")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "ANÁLISE DE BREAKEVEN ENTRE INDEXADORES"
    c.font = Font(bold=True, size=14, color=BRANCO)
    c.fill = hdr_fill(AZUL_ESCURO); c.alignment = centro()
    ws.row_dimensions[1].height = 26

    # Bloco 1: PRE vs IPCA+
    ws["A3"] = "BREAKEVEN PRÉ vs IPCA+  (Equação de Fisher)"; ws["A3"].font = Font(bold=True, size=11)
    ws["A4"] = "Inflação implícita = (1+PRE%)/(1+IPCA+%) − 1"
    ws["A4"].font = Font(italic=True, size=9, color="FF606060")
    aplicar_cabecalho(ws, 5, ["Taxa PRÉ (%)", "Taxa IPCA+ (%)", "Inflação Implícita (% a.a.)",
                               "Interpretação"])
    combos_pre = [
        (10.5, 5.5), (11.0, 5.5), (11.5, 5.5), (11.5, 6.0), (12.0, 6.0),
        (12.5, 6.0), (12.5, 6.5), (13.0, 6.5), (13.5, 6.5), (14.0, 6.5),
    ]
    for i, (taxa_pre, taxa_ipca) in enumerate(combos_pre, 6):
        inf_i = infla_implicita(taxa_pre, taxa_ipca)
        interp = f"IPCA>{inf_i:.2f}% → IPCA+ melhor" if infla > inf_i else f"IPCA<{inf_i:.2f}% → PRÉ melhor"
        fill = hdr_fill(VERDE_CLARO) if infla > inf_i else hdr_fill(AMARELO)
        for col, val, fmt in [
            (1, taxa_pre, '0.00"%"'), (2, taxa_ipca, '0.00"%"'),
            (3, inf_i, '0.0000"%"'), (4, interp, '@'),
        ]:
            c = ws.cell(row=i, column=col, value=val)
            c.number_format = fmt; c.fill = fill; c.border = borda_fina()

    # Bloco 2: %CDI vs PRE
    linha_b2 = 6 + len(combos_pre) + 2
    ws.cell(row=linha_b2, column=1, value="BREAKEVEN %CDI vs PRÉ  (CDI médio de equilíbrio)").font = Font(bold=True, size=11)
    ws.cell(row=linha_b2 + 1, column=1, value="CDI_break = (1+PRE%)^(100/pct) − 1").font = Font(italic=True, size=9, color="FF606060")
    aplicar_cabecalho(ws, linha_b2 + 2, ["Taxa PRÉ (%)", "%CDI", "CDI Breakeven (% a.a.)",
                                           "Atual CDI Proj.", "Situação"])
    combos_pct = [
        (11.5, 95), (11.5, 98), (11.5, 100), (11.5, 105),
        (12.0, 98), (12.0, 100), (12.0, 105), (12.0, 110),
        (12.5, 105), (12.5, 110), (13.0, 110), (13.0, 115),
    ]
    for i, (taxa_pre, pct) in enumerate(combos_pct, linha_b2 + 3):
        cdi_bk = cdi_breakeven_pct(taxa_pre, pct)
        situacao = f"%CDI melhor (CDI>{cdi_bk:.2f}%)" if cdi > cdi_bk else f"PRÉ melhor (CDI<{cdi_bk:.2f}%)"
        fill = hdr_fill(VERDE_CLARO) if cdi > cdi_bk else hdr_fill(AMARELO)
        for col, val, fmt in [
            (1, taxa_pre, '0.00"%"'), (2, pct, '0.00"%"'),
            (3, cdi_bk, '0.0000"%"'), (4, cdi, '0.00"%"'), (5, situacao, '@'),
        ]:
            c = ws.cell(row=i, column=col, value=val)
            c.number_format = fmt; c.fill = fill; c.border = borda_fina()

    set_larguras(ws, {"A": 16, "B": 16, "C": 26, "D": 20, "E": 32, "F": 16})
    return ws


# ── Planilha 5: Formulas (referencia) ────────────────────────────────────────

def sheet_formulas(wb, cdi: float, infla: float):
    ws = wb.create_sheet("Fórmulas e Convenções")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"].value = "FÓRMULAS E CONVENÇÕES — SWAP DE INDEXADORES (Brasil)"
    ws["A1"].font = Font(bold=True, size=13, color=BRANCO)
    ws["A1"].fill = hdr_fill(AZUL_ESCURO); ws["A1"].alignment = centro()
    ws.row_dimensions[1].height = 26

    linhas = [
        ("", ""),
        ("CADEIA DE EQUIVALÊNCIA", ""),
        ("%CDI  ──A──▶  CDI+spread  ──B──▶  PRÉ  ──C──▶  IPCA+spread", ""),
        ("", ""),
        ("PASSO A — %CDI ↔ CDI+spread", ""),
        ("spread_cdi = [(1+CDI/100)^(pct/100−1) − 1] × 100", "Independente do prazo. Confirmado ANBIMA/B3."),
        ("pct_cdi    = [1 + ln(1+spread/100) / ln(1+CDI/100)] × 100", "Inversa da fórmula acima."),
        (f"Exemplo: 98% CDI com CDI={cdi:.1f}%  →  spread ≈ {pct_para_cdi(98, cdi):.4f}% a.a.", ""),
        ("", ""),
        ("PASSO C — CDI+spread ↔ IPCA+spread  (Fisher Multiplicativo)", ""),
        ("leg_cdi_nominal = (1+CDI/100) × (1+spread_cdi/100)", "Capitalização multiplicativa (exata)."),
        ("spread_ipca = [leg_nominal / (1+IPCA/100) − 1] × 100", "Equação de Fisher. Independe do prazo."),
        ("Inverso: spread_cdi = [(1+spread_ipca/100)×(1+IPCA/100)/(1+CDI/100) − 1] × 100", ""),
        ("", ""),
        ("PASSO DIRETO — %CDI ↔ IPCA+ (encadeia A+C)", ""),
        ("spread_ipca = [(1+CDI/100)^(pct/100) / (1+IPCA/100) − 1] × 100", ""),
        ("pct_cdi = ln[(1+spread_ipca/100)×(1+IPCA/100)] / ln(1+CDI/100) × 100", ""),
        ("", ""),
        ("BREAKEVEN PRÉ vs IPCA+  (Fisher)", ""),
        ("infla_implícita = (1+PRE/100)/(1+IPCA_real/100) − 1  ×100  [% a.a.]", ""),
        ("Interpretação: se IPCA esperado > infla_implícita → IPCA+ preferível.", ""),
        ("", ""),
        ("BREAKEVEN %CDI vs PRÉ", ""),
        ("CDI_break = [(1+PRE/100)^(100/pct) − 1] × 100  [% a.a.]", ""),
        ("Interpretação: se CDI médio esperado > CDI_break → %CDI preferível.", ""),
        ("", ""),
        ("NII POR PERÍODO (Net Interest Income)", ""),
        ("NII = Nocional × [(1+taxa_aa/100)^(du/252) − 1]", "du = dias úteis do período."),
        ("Para leg IPCA: taxa_aa ≈ spread_ipca + IPCA_projetado (simplificado)", "Forma exata: VNA_fim/VNA_ini × (1+spread)^(du/252)"),
        ("", ""),
        ("CONVENÇÕES DO MERCADO BRASILEIRO", ""),
        ("Base de tempo: 252 dias úteis (B3/ANBIMA)", ""),
        ("CDI: over diário (252 dias úteis). Expresso como % a.a.", ""),
        ("IPCA: mensal. VNA corrige pelo índice acumulado.", ""),
        ("Spread aditivo vs multiplicativo: debentures usam spread multiplicativo", "Diferença relevante para spreads > 2%."),
        ("Fonte da curva DI: _curva_di.csv (sync_b3_curve.py / importar_curva_historica.py)", ""),
        ("Fonte da curva IPCA real: ANBIMA ETTJ-IPCA ou NTN-B de prazo similar (input manual)", ""),
    ]
    for i, (col_a, col_b) in enumerate(linhas, 3):
        c_a = ws.cell(row=i, column=1, value=col_a)
        c_b = ws.cell(row=i, column=4, value=col_b)
        if col_a.isupper() and col_b == "":
            c_a.font = Font(bold=True, size=10, color=AZUL_ESCURO)
        elif col_a.startswith("  ") or "→" in col_a or "=" in col_a or "^" in col_a:
            c_a.font = Font(name="Courier New", size=9)
        elif col_b:
            c_b.font = Font(italic=True, size=9, color="FF606060")

    ws.merge_cells("A1:D1")
    ws.column_dimensions["A"].width = 75
    ws.column_dimensions["D"].width = 45
    return ws


# ── Main ─────────────────────────────────────────────────────────────────────

def parse_args():
    import argparse
    p = argparse.ArgumentParser(description="Gera planilha de swap de indexadores")
    p.add_argument("--cdi",        type=float, default=13.50, help="CDI projetado %% a.a. (default 13.50)")
    p.add_argument("--ipca",       type=float, default=5.20,  help="IPCA projetado %% a.a. (default 5.20)")
    p.add_argument("--spread-cdi", type=float, default=2.00,  help="Spread CDI+ em %% a.a. (default 2.00)")
    p.add_argument("--pct-cdi",    type=float, default=100.0, help="%%CDI (default 100)")
    p.add_argument("--nocional",   type=float, default=1_000_000, help="Nocional R$ (default 1M)")
    p.add_argument("--inicio",     type=str,   default=None,  help="Data inicio YYYY-MM-DD (default hoje)")
    p.add_argument("--fim",        type=str,   default=None,  help="Data fim YYYY-MM-DD (default +5 anos)")
    p.add_argument("--output",     type=str,   default=None,  help="Caminho do arquivo de saida")
    return p.parse_args()


def main():
    args = parse_args()
    cdi       = args.cdi
    infla     = args.ipca
    spread_cdi = args.spread_cdi
    pct_cdi   = args.pct_cdi
    nocional  = args.nocional

    hoje = datetime.date.today()
    data_ini = datetime.date.fromisoformat(args.inicio) if args.inicio else hoje
    data_fim = datetime.date.fromisoformat(args.fim) if args.fim else \
        datetime.date(hoje.year + 5, hoje.month, hoje.day)

    out_path = args.output
    if not out_path:
        nome = f"swap_{hoje.strftime('%Y%m%d')}_CDI{cdi:.0f}_IPCA{infla:.0f}.xlsx"
        out_path = str(Path(__file__).resolve().parent.parent / "data" / nome)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove Sheet vazia default

    print(f"Gerando planilha de swap...")
    print(f"  CDI proj = {cdi:.2f}%  |  IPCA proj = {infla:.2f}%")
    print(f"  CDI+{spread_cdi:.2f}%  ->  IPCA+{cdi_para_ipca(spread_cdi, cdi, infla):.4f}%")
    print(f"  {pct_cdi:.0f}%CDI  ->  CDI+{pct_para_cdi(pct_cdi,cdi):.4f}%  ->  IPCA+{pct_para_ipca(pct_cdi,cdi,infla):.4f}%")

    sheet_equivalencias(wb, cdi, infla)
    sheet_swap_cdi_ipca(wb, cdi, infla, spread_cdi, nocional, data_ini, data_fim)
    sheet_swap_pct_ipca(wb, cdi, infla, pct_cdi, nocional, data_ini, data_fim)
    sheet_breakeven(wb, cdi, infla)
    sheet_formulas(wb, cdi, infla)

    wb.save(out_path)
    print(f"  Salvo em: {out_path}")
    print("  Abas: Equivalencias | Swap CDI+<>IPCA+ | Swap %CDI<>IPCA+ | Breakeven | Formulas")


if __name__ == "__main__":
    main()
