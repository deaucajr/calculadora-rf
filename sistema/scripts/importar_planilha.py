"""
Le o CADASTRO_ATIVO.xlsx preenchido e gera o CSV em fluxos_manual/.

Uso:
  python scripts/importar_planilha.py CADASTRO_ATIVO.xlsx
  python scripts/importar_planilha.py meu_ativo.xlsx --ver     # mostra o CSV sem salvar

Formula usada:
  IPCA+ / PRE  ->  FC%_juros  = frac_restante x [(1+taxa)^(du_periodo/252) - 1]
                   FC%_amort  = amort_pct / 100
  CDI+ / %CDI  ->  nao suportado no manual (use importar_fluxos.py com a API B3)
"""
import sys
from datetime import date, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

try:
    import openpyxl
except ImportError:
    print("ERRO: pip install openpyxl"); sys.exit(1)

from src.paths import fluxos_manual_dir

SEP = "\t"


# ── Dias uteis ────────────────────────────────────────────────────────────────

def _carregar_feriados() -> set:
    try:
        p = ROOT / "data" / "fluxos" / "_feriados.csv"
        if not p.exists():
            p = ROOT / "data" / "fluxos" / "_feriados.csv"
        if p.exists():
            return {date.fromisoformat(ln.strip()) for ln in p.read_text().splitlines() if ln.strip()}
    except Exception:
        pass
    return set()


_FERIADOS = None


def feriados() -> set:
    global _FERIADOS
    if _FERIADOS is None:
        _FERIADOS = _carregar_feriados()
    return _FERIADOS


def du(d0: date, d1: date) -> int:
    """Dias uteis de d0 (exclusive) ate d1 (inclusive)."""
    if d1 <= d0:
        return 0
    fer = feriados()
    count = 0
    d = d0 + timedelta(1)
    while d <= d1:
        if d.weekday() < 5 and d not in fer:
            count += 1
        d += timedelta(1)
    return count


# ── Leitura do Excel ──────────────────────────────────────────────────────────

def _ler_data(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, date):
        return v if not hasattr(v, "hour") else v.date()
    if hasattr(v, "date"):
        return v.date()
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _ler_float(v, default=None) -> float | None:
    if v is None:
        return default
    try:
        return float(str(v).replace(",", ".").strip())
    except (ValueError, AttributeError):
        return default


def _ler_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def ler_excel(caminho: str) -> tuple[dict, list]:
    wb = openpyxl.load_workbook(caminho, data_only=True)

    # ── Aba Identificacao ──────────────────────────────────────────────────────
    aba_id = None
    for nome in wb.sheetnames:
        if "identifica" in nome.lower() or "ident" in nome.lower():
            aba_id = wb[nome]
            break
    if aba_id is None:
        aba_id = wb.worksheets[0]

    # Le pares (linha da coluna A, valor da coluna B), ignora cabecalho e dicas
    meta_raw = {}
    for row in aba_id.iter_rows(min_row=3, max_col=2, values_only=True):
        chave, valor = row
        if chave and valor is not None and str(valor).strip():
            chave_norm = (str(chave).strip()
                          .upper()
                          .replace("*", "")
                          .strip()
                          .replace(" ", "_")
                          .replace("(R$)", "")
                          .replace("(%_A.A.)", "")
                          .strip("_"))
            meta_raw[chave_norm] = valor

    # ── Aba Fluxo_de_Caixa ────────────────────────────────────────────────────
    aba_fl = None
    for nome in wb.sheetnames:
        if "fluxo" in nome.lower():
            aba_fl = wb[nome]
            break
    if aba_fl is None:
        aba_fl = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]

    linhas_fluxo = []
    for row in aba_fl.iter_rows(min_row=4, max_col=4, values_only=True):
        data_v, juros_v, amort_v, taxa_v = row
        d = _ler_data(data_v)
        if d is None:
            continue
        juros_str = _ler_str(juros_v).upper()
        paga_juros = juros_str in ("S", "SIM", "Y", "YES", "X", "1", "TRUE")
        amort_pct  = _ler_float(amort_v, 0.0)
        taxa       = _ler_float(taxa_v, None)   # None = usa taxa_emissao
        linhas_fluxo.append((d, paga_juros, amort_pct, taxa))

    return meta_raw, sorted(linhas_fluxo, key=lambda x: x[0])


# ── Validacao ─────────────────────────────────────────────────────────────────

INDEXADORES_VALIDOS = {"IPCA+", "IPCA", "PRE", "PRÉ", "CDI+", "%CDI", "PCTCDI"}


def validar(meta: dict, linhas: list) -> list[str]:
    erros = []

    # Campos obrigatorios
    ticker = _ler_str(meta.get("TICKER") or meta.get("TICKER_"))
    if not ticker:
        erros.append("Ticker nao preenchido na aba Identificacao.")

    idx_raw = _ler_str(meta.get("INDEXADOR") or "").upper().replace("+", "+").strip()
    if not idx_raw:
        erros.append("Indexador nao preenchido.")
    elif "CDI" in idx_raw and "IPCA" not in idx_raw:
        erros.append(
            f"Indexador '{idx_raw}' (CDI+/%CDI) nao suportado no cadastro manual. "
            "Use importar_fluxos.py com a API B3."
        )

    taxa_emissao = _ler_float(meta.get("TAXA_DE_EMISSAO") or meta.get("TAXA_EMISSAO") or
                               meta.get("TAXA") or meta.get("TAXA_DE_EMISSAO_"))
    if taxa_emissao is None or taxa_emissao <= 0:
        erros.append("Taxa de Emissao nao preenchida ou invalida.")

    d_emissao = _ler_data(meta.get("DATA_DE_EMISSAO") or meta.get("DATA_EMISSAO") or
                           meta.get("DATA_DE_EMISSAO_"))
    if d_emissao is None:
        erros.append("Data de Emissao nao preenchida ou invalida.")

    vencimento = _ler_data(meta.get("VENCIMENTO"))
    if vencimento is None:
        erros.append("Vencimento nao preenchido ou invalido.")

    # Fluxo
    if not linhas:
        erros.append("Nenhuma linha de fluxo encontrada na aba Fluxo_de_Caixa.")
        return erros

    amort_total = sum(a for _, _, a, _ in linhas if a)
    if amort_total > 0 and abs(amort_total - 100.0) > 0.01:
        erros.append(
            f"Soma das amortizacoes = {amort_total:.2f}% (deveria ser 0 para bullet ou 100 para amortizante)."
        )

    if vencimento and linhas:
        ultima_data = linhas[-1][0]
        if ultima_data != vencimento:
            erros.append(
                f"Ultima data do fluxo ({ultima_data}) diferente do Vencimento ({vencimento}). "
                "Verifique se o vencimento esta correto."
            )

    return erros


# ── Calculo do FC% ────────────────────────────────────────────────────────────

def calcular_csv(meta: dict, linhas: list) -> str:
    # Normaliza metadados
    ticker      = _ler_str(meta.get("TICKER") or meta.get("TICKER_")).upper()
    tipo_raw    = _ler_str(meta.get("TIPO") or "DEB").upper()
    tipo        = tipo_raw if tipo_raw in ("DEB", "CRI", "CRA") else "DEB"
    idx_raw     = _ler_str(meta.get("INDEXADOR") or "IPCA+").upper()
    indexador   = "IPCA" if "IPCA" in idx_raw else "PRE"
    emissor     = _ler_str(meta.get("EMISSOR") or "MANUAL").replace(SEP, " ")
    vne         = _ler_float(meta.get("VNE") or meta.get("VNE_") or 1000.0, 1000.0)
    taxa_em     = _ler_float(meta.get("TAXA_DE_EMISSAO") or meta.get("TAXA_EMISSAO") or
                              meta.get("TAXA") or meta.get("TAXA_DE_EMISSAO_"), 0.0)
    rating      = _ler_str(meta.get("RATING") or "")
    garantia    = _ler_str(meta.get("GARANTIA") or "")
    d_emissao   = _ler_data(meta.get("DATA_DE_EMISSAO") or meta.get("DATA_EMISSAO") or
                              meta.get("DATA_DE_EMISSAO_"))
    vencimento  = _ler_data(meta.get("VENCIMENTO"))
    hoje        = date.today().isoformat()

    # META
    out = []
    out.append(SEP.join(["META", "TICKER",       ticker]))
    out.append(SEP.join(["META", "TIPO",          tipo]))
    out.append(SEP.join(["META", "INDEXADOR",     indexador]))
    out.append(SEP.join(["META", "EMISSOR",       emissor]))
    out.append(SEP.join(["META", "METHOD",        indexador]))
    out.append(SEP.join(["META", "INICIO_RENT",   d_emissao.isoformat() if d_emissao else ""]))
    out.append(SEP.join(["META", "VENCIMENTO",    vencimento.isoformat() if vencimento else ""]))
    out.append(SEP.join(["META", "VNE",           f"{vne:.2f}"]))
    out.append(SEP.join(["META", "TAXA_EMISSAO",  f"{taxa_em:.4f}"]))
    out.append(SEP.join(["META", "TAXA_REF",      f"{taxa_em:.4f}"]))
    out.append(SEP.join(["META", "DATA_FLUXO",    hoje]))
    if rating:
        out.append(SEP.join(["META", "RATING",    rating]))
    if garantia:
        out.append(SEP.join(["META", "GARANTIA",  garantia]))
    out.append(SEP.join(["META", "FONTE",         "MANUAL"]))

    # FLUXOS: calcula FC%
    # Bullet: amort_total = 0 em todas as linhas exceto vencimento
    amort_total = sum(a for _, _, a, _ in linhas if a)
    is_bullet = amort_total == 0.0 or (len(linhas) == 1)

    d_ref = d_emissao or linhas[0][0]  # data de referencia para du_periodo
    cumul_amort = 0.0                   # % acumulado (0-100)
    amortpct_linhas = []

    for data, paga_juros, amort_pct, taxa_linha in linhas:
        taxa = taxa_linha if (taxa_linha is not None and taxa_linha > 0) else taxa_em
        remaining = (100.0 - cumul_amort) / 100.0

        du_periodo = du(d_ref, data)          # du do periodo de juros
        du_total   = du(d_emissao or data, data) if d_emissao else du_periodo

        # Linha de JUROS
        if paga_juros and remaining > 0 and du_periodo > 0:
            fc_j = remaining * ((1 + taxa / 100) ** (du_periodo / 252) - 1)
            vf_j = fc_j * vne
            pv_j = vf_j / (1 + taxa / 100) ** (du_total / 252) if du_total > 0 else vf_j
            out.append(SEP.join([
                "FLUXO", data.isoformat(), "J",
                f"{vf_j:.6f}", f"{pv_j:.6f}",
                str(du_total), f"{fc_j:.10f}",
            ]))

        # Linha de AMORTIZACAO
        if amort_pct and amort_pct > 0:
            fc_a = amort_pct / 100.0
            vf_a = fc_a * vne
            pv_a = vf_a / (1 + taxa / 100) ** (du_total / 252) if du_total > 0 else vf_a
            out.append(SEP.join([
                "FLUXO", data.isoformat(), "A",
                f"{vf_a:.6f}", f"{pv_a:.6f}",
                str(du_total), f"{fc_a:.10f}",
            ]))
            amortpct_linhas.append((data, amort_pct))
            cumul_amort += amort_pct

        # Bullet: ao chegar no vencimento, emite linha A de amortizacao
        if is_bullet and data == vencimento and amort_pct == 0:
            fc_a = 1.0
            vf_a = vne
            pv_a = vf_a / (1 + taxa / 100) ** (du_total / 252) if du_total > 0 else vf_a
            out.append(SEP.join([
                "FLUXO", data.isoformat(), "A",
                f"{vf_a:.6f}", f"{pv_a:.6f}",
                str(du_total), "1.0000000000",
            ]))
            amortpct_linhas.append((data, 100.0))

        if paga_juros or (amort_pct and amort_pct > 0) or (is_bullet and data == vencimento):
            d_ref = data

    # VNA — ponto de ancoragem
    out.append(SEP.join(["VNA", hoje,                       f"{vne:.6f}"]))
    if vencimento:
        out.append(SEP.join(["VNA", vencimento.isoformat(), f"{vne:.6f}"]))

    # AMORTPCT (para IPCA amortizante — correcao de VNA em datas intermediarias)
    if indexador == "IPCA" and amortpct_linhas:
        for d_a, pct_a in amortpct_linhas:
            if abs(pct_a - 100.0) > 0.01:   # nao emite linha p/ bullet (100% no venc)
                out.append(SEP.join(["AMORTPCT", d_a.isoformat(), f"{pct_a:.8f}"]))

    return "\r\n".join(out) + "\r\n"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    if not args or args[0].startswith("--"):
        print("Uso: python scripts/importar_planilha.py ARQUIVO.xlsx [--ver]")
        print("     --ver   mostra o CSV gerado sem salvar")
        sys.exit(1)

    caminho = args[0]
    apenas_ver = "--ver" in args

    print(f"Lendo {caminho}...")
    try:
        meta, linhas = ler_excel(caminho)
    except Exception as e:
        print(f"ERRO ao ler o Excel: {e}")
        sys.exit(1)

    print(f"  {len(linhas)} linhas de fluxo encontradas")

    erros = validar(meta, linhas)
    if erros:
        print("\nERROS DE VALIDACAO:")
        for e in erros:
            print(f"  - {e}")
        sys.exit(1)

    ticker = _ler_str(meta.get("TICKER") or meta.get("TICKER_")).upper()
    csv_txt = calcular_csv(meta, linhas)

    if apenas_ver:
        print("\n--- CSV gerado ---")
        print(csv_txt)
        return

    destino = fluxos_manual_dir() / f"{ticker}.csv"
    destino.parent.mkdir(parents=True, exist_ok=True)
    with open(destino, "w", encoding="utf-8", newline="") as f:
        f.write(csv_txt)

    print(f"\nOK: {destino}")
    print(f"   Teste no Excel: =RF_PU(\"{ticker}\"; [taxa]; HOJE())")
    print(f"   Fluxo:          =RF_FLUXO(\"{ticker}\"; [taxa]; HOJE())")


if __name__ == "__main__":
    main()
